from selenium import webdriver
import time
from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == "__main__":
    driver = webdriver.Chrome('./chromedriver')
    driver.maximize_window()
    driver.get('https://www.sketchymedical.com/#!/login')
    time.sleep(1)
    #ID = input("아이디를 입력하세요::")
    ID = 'ghdwn600@gmail.com'
    #PW = input("비밀번호 입력하세요::")
    PW = '1234567890'

    driver.find_element_by_xpath('//*[@id="login-page"]/div/form/div/div[1]/input[1]').send_keys(ID)
    driver.find_element_by_xpath('//*[@id="login-page"]/div/form/div/div[1]/input[2]').send_keys(PW)
    driver.find_element_by_xpath('//*[@id="login-form-btn"]').click()
    # dummy xlsx 저장
    wb  = Workbook()
    ws1 = wb.active
    wb.save('dummy.xlsx')
    isFirst = True
    while True:
        input("페이지를 띄운후 엔터를 쳐주세요::")
        title = driver.find_element_by_xpath('/html/body/div[1]/div/div/div/div[1]/h3')
        print("제목",title.text)
        btns = driver.find_elements_by_id('hotspot-position')
        idx = 1
        if isFirst:
            wb = load_workbook('dummy.xlsx')
            ws1 = wb.active
            ws1.append([title.text])
            isFirst = False
        else:
            wb = load_workbook('data.xlsx')
            ws1 = wb.active
            ws1.append([''])
            ws1.append([title.text])
        for btn in btns:
            while True:
                try:
                    btn.click()
                    time.sleep(0.5)
                    ws1.append([str(idx)+'. '+driver.find_element_by_id('tooltip-hostspot-content').text])
                    print(str(idx)+'. '+driver.find_element_by_id('tooltip-hostspot-content').text)
                    title.click()
                    time.sleep(2)
                    break
                except:  # 가져오는 부분에난 에러
                    driver.find_element_by_xpath('/html/body/div[1]/div/div/div/div[1]/span/h3').click()
                    time.sleep(2)
                    break
            idx += 1
        print(">>> 저장중...")
        wb.save('data.xlsx')
    driver.quit()